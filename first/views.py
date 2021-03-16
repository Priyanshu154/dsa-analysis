from django.shortcuts import render, redirect
from django.contrib.auth import forms, authenticate, login, logout

from django.contrib.auth.decorators import login_required
from rest_framework.response import Response
from rest_framework.decorators import api_view
from .models import cp_coder
import os
import openpyxl as xl
from openpyxl.utils import column_index_from_string
from django.core.exceptions import ValidationError
import json

@login_required( login_url='login' )
def index(request):
    print( 'here' )
    return render(request,'build/index.html')

# Create your views here.

workpath = os.path.dirname(os.path.abspath(__file__))

xx = os.path.join(workpath,'problems.xlsx' )
wb = xl.load_workbook(xx, data_only=True)
sheet1 = wb['Sheet1']


def logout_view(request):
    logout(request)
    return redirect( 'register' )


def register_view(request):
    #register the user
    if( request.method != 'POST' ):  return render( request, 'build/index.html' )
    print( request.POST.get( 'password1' ) )
    form = forms.UserCreationForm(request.POST)
    if form.is_valid():
        user =form.save()
        #use the below code when model has a new column
        #new_coder = cp_coder( user=user,name=request.data.get('naam') )
        #new_coder.save()
        login( request,user )
        return redirect('codeforces_handle')
    else:
        return render( request, 'build/index.html',{ 'form': form.errors.as_json() } )

def login_view(request):
    form = forms.AuthenticationForm(request.POST)
    if request.method == "POST":
        user = authenticate(request, username= request.POST.get('username'), password= request.POST.get('password'))
        if user is not None:
            login(request, user)
            print( 'user logged in' )
            return redirect('codeforces_handle')
        else:
            print( form.errors )
            return render( request,'build/index.html',{'form': """ { "username": [ {"message": "username or password is incorrect"}] }  """ } )
    else:
        print( form.errors )
        return render( request, 'build/index.html')

#API
@api_view( ['POST'] )
def get_problems(request):
    tags = request.data[  'tags'  ]
    max_que = request.data[ 'maxQue' ]
    user_rating = request.data[ 'rating' ]
    high = request.data[ 'range' ]
    res = {}

    for tag in tags:
        res[tag] = { 'name' : [], 'url' : [] }

    for tag in tags:
        count = 0
        for i in range(2,sheet1.max_row+1 ):
            if count >= max_que: break

            url = sheet1.cell(i, column_index_from_string('A')).value
            name = sheet1.cell(i, column_index_from_string('B')).value
            rating = sheet1.cell(i, column_index_from_string('C')).value
            sheet_tag = sheet1.cell( i,column_index_from_string('D') ).value

            if sheet_tag == tag:
                if rating == -1:
                    res[tag]['url'].append( url )
                    res[tag]['name'].append( name )
                    count = count + 1
                if rating >= user_rating and rating <= user_rating + high:
                    res[tag]['url'].append( url )
                    res[tag]['name'].append( name )
                    count = count + 1
    arr = []
    for tag in res:
        arr.append( {'tag' : tag } )
        arr[ len(arr)-1 ]['name'] = res[tag]['name']
        arr[len(arr)-1]['url'] = res[tag]['url']
    return Response( arr )

